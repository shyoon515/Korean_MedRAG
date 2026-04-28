#!/usr/bin/env python3
"""
Generate a multi-level hybrid retrieval benchmark report.

This script evaluates:
- Sparse only
- Dense only
- Hybrid alpha sweep from 0.0 to 1.0 in 0.1 steps
- RRF
- Pilot hybrid with deterministic bootstrapped alpha tuning

It reports both Recall@10 and NDCG@10 for:
- per-file analysis
- pooled group analysis

The report is written as markdown, and a machine-readable JSON companion is
written alongside it.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import random
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List, Optional, Sequence


TOP_K = 10
ALPHA_VALUES = [round(idx / 10, 1) for idx in range(11)]
PILOT_SAMPLE_SIZE = 300
PILOT_TOP_ALPHA_COUNT = 3
PILOT_MIN_QUERIES = 600
DEFAULT_REPORT_MD = "docs/hybrid_retrieval_benchmark_report.md"
DEFAULT_REPORT_JSON = "eval_result_ret/hybrid_retrieval_benchmark_summary.json"
DEFAULT_REPORT_XLSX = "eval_result_ret/hybrid_retrieval_benchmark_summary.xlsx"


def load_json(path: Path) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def normalize_scores(scores: Sequence[float]) -> List[float]:
    if not scores:
        return []
    min_score = min(scores)
    max_score = max(scores)
    if min_score == max_score:
        return [1.0] * len(scores)
    return [(score - min_score) / (max_score - min_score) for score in scores]


def dcg_from_labels(labels: Sequence[int], k: int = TOP_K) -> float:
    total = 0.0
    for idx, label in enumerate(labels[:k], start=1):
        if label > 0:
            total += label / math.log2(idx + 1)
    return total


def ndcg_from_labels(labels: Sequence[int], ideal_dcg: float, k: int = TOP_K) -> float:
    if ideal_dcg <= 0:
        return 0.0
    return dcg_from_labels(labels, k=k) / ideal_dcg


def mean_or_zero(values: Sequence[float]) -> float:
    return mean(values) if values else 0.0


def build_alpha_values(step: float) -> List[float]:
    if step <= 0:
        raise ValueError("alpha step must be > 0")

    steps = int(round(1.0 / step))
    values = [round(min(idx * step, 1.0), 10) for idx in range(steps + 1)]
    if values[-1] != 1.0:
        values.append(1.0)
    return values


def stable_seed(seed: int, salt: str) -> int:
    digest = hashlib.sha1(salt.encode("utf-8")).hexdigest()
    return seed + int(digest[:12], 16)


def discover_cache_files(cache_root: Path) -> List[str]:
    files = []
    for path in sorted(cache_root.glob("*.json")):
        if path.name == "manifest.json" or path.name.endswith("_manifest.json"):
            continue
        files.append(path.name)
    return files


def filter_file_names(
    file_names: Sequence[str],
    include_files: Optional[Sequence[str]] = None,
    max_files: int = 0,
) -> List[str]:
    selected = list(file_names)
    if include_files:
        include_set = {name.strip() for name in include_files if name.strip()}
        selected = [name for name in selected if name in include_set]
    if max_files > 0:
        selected = selected[:max_files]
    return selected


def group_name_for_file(file_name: str) -> Optional[str]:
    if file_name.startswith(("TL_", "VL_")):
        return "AIHub TL+VL"
    if file_name.startswith("KorMedMCQA_dentist_"):
        return "KorMedMCQA_dentist"
    if file_name.startswith("KorMedMCQA_doctor_"):
        return "KorMedMCQA_doctor"
    return None


def aihub_dataset_name_for_file(file_name: str) -> Optional[str]:
    if not file_name.startswith(("TL_", "VL_")):
        return None
    base_name = file_name.split("_", 1)[1]
    if base_name.endswith(".json"):
        base_name = base_name[:-5]
    return base_name


def extract_label_map(judge_entry: Dict[str, Any]) -> Dict[str, int]:
    label_map: Dict[str, int] = {}
    for item in judge_entry.get("retrieved_items", []):
        chunk_id = str(item.get("chunk_id", "")).strip()
        relevance = item.get("llm_relevance_01")
        if chunk_id and relevance is not None:
            label_map[chunk_id] = max(label_map.get(chunk_id, 0), int(relevance))
    return label_map


def combine_label_maps(*maps: Dict[str, int]) -> Dict[str, int]:
    combined: Dict[str, int] = {}
    for label_map in maps:
        for chunk_id, relevance in label_map.items():
            combined[chunk_id] = max(combined.get(chunk_id, 0), int(relevance))
    return combined


def combine_chunk_ids(*ordered_lists: Sequence[str]) -> List[str]:
    seen = set()
    merged: List[str] = []
    for ordered in ordered_lists:
        for chunk_id in ordered:
            if chunk_id not in seen:
                seen.add(chunk_id)
                merged.append(chunk_id)
    return merged


def load_retrieval_results(cache_path: Path, file_name: str) -> Dict[str, Any]:
    file_path = cache_path / file_name
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_source_view(retrieval_entry: Dict[str, Any], judge_entry: Dict[str, Any], top_k: int = TOP_K) -> Dict[str, Any]:
    retrieved_items = list(retrieval_entry.get("retrieved_items", []))[:top_k]
    raw_scores = [float(item.get("score", 0.0)) for item in retrieved_items]
    norm_scores = normalize_scores(raw_scores)
    label_map = extract_label_map(judge_entry)

    ordered_chunk_ids: List[str] = []
    raw_score_map: Dict[str, float] = {}
    norm_score_map: Dict[str, float] = {}
    rank_map: Dict[str, int] = {}
    top_labels: List[int] = []

    for rank, (item, norm_score) in enumerate(zip(retrieved_items, norm_scores), start=1):
        chunk_id = str(item.get("chunk_id", "")).strip()
        if not chunk_id:
            continue
        ordered_chunk_ids.append(chunk_id)
        raw_score_map[chunk_id] = float(item.get("score", 0.0))
        norm_score_map[chunk_id] = norm_score
        rank_map[chunk_id] = rank
        top_labels.append(int(label_map.get(chunk_id, 0)))

    ideal_dcg = dcg_from_labels(sorted(label_map.values(), reverse=True), k=top_k)

    return {
        "ordered_chunk_ids": ordered_chunk_ids,
        "raw_score_map": raw_score_map,
        "norm_score_map": norm_score_map,
        "rank_map": rank_map,
        "label_map": label_map,
        "top_labels": top_labels,
        "ideal_dcg": ideal_dcg,
    }


def build_query_record(
    file_name: str,
    query_idx: int,
    dense_entry: Dict[str, Any],
    bm25_entry: Dict[str, Any],
    dense_judge_entry: Dict[str, Any],
    bm25_judge_entry: Dict[str, Any],
) -> Dict[str, Any]:
    dense_view = build_source_view(dense_entry, dense_judge_entry)
    bm25_view = build_source_view(bm25_entry, bm25_judge_entry)
    label_map = combine_label_maps(dense_view["label_map"], bm25_view["label_map"])
    ideal_dcg = dcg_from_labels(sorted(label_map.values(), reverse=True), k=TOP_K)

    return {
        "file_name": file_name,
        "query_idx": query_idx,
        "dense": dense_view,
        "bm25": bm25_view,
        "label_map": label_map,
        "ideal_dcg": ideal_dcg,
        "candidate_ids": combine_chunk_ids(
            dense_view["ordered_chunk_ids"],
            bm25_view["ordered_chunk_ids"],
        ),
    }


def build_query_records_for_file(
    file_name: str,
    dense_cache: Dict[str, Any],
    bm25_cache: Dict[str, Any],
    dense_judge_cache: Dict[str, Any],
    bm25_judge_cache: Dict[str, Any],
) -> List[Dict[str, Any]]:
    dense_entries = dense_cache.get("entries", [])
    bm25_entries = bm25_cache.get("entries", [])
    dense_judge_entries = dense_judge_cache.get("entries", [])
    bm25_judge_entries = bm25_judge_cache.get("entries", [])

    limit = min(
        len(dense_entries),
        len(bm25_entries),
        len(dense_judge_entries),
        len(bm25_judge_entries),
    )

    records: List[Dict[str, Any]] = []
    for query_idx in range(limit):
        dense_entry = dense_entries[query_idx]
        bm25_entry = bm25_entries[query_idx]
        if not dense_entry.get("retrieved_items") or not bm25_entry.get("retrieved_items"):
            continue

        record = build_query_record(
            file_name=file_name,
            query_idx=query_idx,
            dense_entry=dense_entry,
            bm25_entry=bm25_entry,
            dense_judge_entry=dense_judge_entries[query_idx],
            bm25_judge_entry=bm25_judge_entries[query_idx],
        )
        records.append(record)

    return records


def query_metrics(top_chunk_ids: Sequence[str], label_map: Dict[str, int], ideal_dcg: float) -> Dict[str, float]:
    labels = [int(label_map.get(chunk_id, 0)) for chunk_id in top_chunk_ids[:TOP_K]]
    recall_at_10 = sum(labels) / float(TOP_K)
    ndcg_at_10 = ndcg_from_labels(labels, ideal_dcg, k=TOP_K)
    return {
        "recall_at_10": recall_at_10,
        "ndcg_at_10": ndcg_at_10,
        "relevant_count": float(sum(labels)),
    }


def rank_dense(record: Dict[str, Any]) -> List[str]:
    return list(record["dense"]["ordered_chunk_ids"][:TOP_K])


def rank_sparse(record: Dict[str, Any]) -> List[str]:
    return list(record["bm25"]["ordered_chunk_ids"][:TOP_K])


def rank_rrf(record: Dict[str, Any]) -> List[str]:
    dense_rank_map = record["dense"]["rank_map"]
    bm25_rank_map = record["bm25"]["rank_map"]
    candidate_ids = record["candidate_ids"]

    scored = []
    for chunk_id in candidate_ids:
        score = 0.0
        if chunk_id in dense_rank_map:
            score += 1.0 / (1.0 + dense_rank_map[chunk_id])
        if chunk_id in bm25_rank_map:
            score += 1.0 / (1.0 + bm25_rank_map[chunk_id])
        scored.append(
            (
                chunk_id,
                score,
                dense_rank_map.get(chunk_id, 999),
                bm25_rank_map.get(chunk_id, 999),
            )
        )

    scored.sort(key=lambda item: (-item[1], item[2], item[3], item[0]))
    return [chunk_id for chunk_id, *_ in scored[:TOP_K]]


def rank_hybrid(record: Dict[str, Any], alpha: float) -> List[str]:
    dense_norm_map = record["dense"]["norm_score_map"]
    bm25_norm_map = record["bm25"]["norm_score_map"]
    dense_rank_map = record["dense"]["rank_map"]
    bm25_rank_map = record["bm25"]["rank_map"]
    candidate_ids = record["candidate_ids"]

    scored = []
    for chunk_id in candidate_ids:
        score = alpha * dense_norm_map.get(chunk_id, 0.0) + (1.0 - alpha) * bm25_norm_map.get(chunk_id, 0.0)
        scored.append(
            (
                chunk_id,
                score,
                dense_rank_map.get(chunk_id, 999),
                bm25_rank_map.get(chunk_id, 999),
            )
        )

    scored.sort(key=lambda item: (-item[1], item[2], item[3], item[0]))
    return [chunk_id for chunk_id, *_ in scored[:TOP_K]]


def evaluate_fixed_method(records: Sequence[Dict[str, Any]], method_name: str) -> Dict[str, Any]:
    recall_scores: List[float] = []
    ndcg_scores: List[float] = []

    for record in records:
        if method_name == "dense_only":
            ranked_chunk_ids = rank_dense(record)
        elif method_name == "sparse_only":
            ranked_chunk_ids = rank_sparse(record)
        elif method_name == "rrf":
            ranked_chunk_ids = rank_rrf(record)
        else:
            raise ValueError(f"Unknown fixed method: {method_name}")

        metrics = query_metrics(ranked_chunk_ids, record["label_map"], record["ideal_dcg"])
        recall_scores.append(metrics["recall_at_10"])
        ndcg_scores.append(metrics["ndcg_at_10"])

    return {
        "recall_at_10": mean_or_zero(recall_scores),
        "ndcg_at_10": mean_or_zero(ndcg_scores),
        "n_queries": len(records),
        "per_query": {
            "recall_at_10": recall_scores,
            "ndcg_at_10": ndcg_scores,
        },
    }


def evaluate_alpha_curve(records: Sequence[Dict[str, Any]], alpha_values: Sequence[float]) -> Dict[float, Dict[str, Any]]:
    curve: Dict[float, Dict[str, List[float]]] = {
        alpha: {"recall_at_10": [], "ndcg_at_10": []} for alpha in alpha_values
    }

    for record in records:
        for alpha in alpha_values:
            ranked_chunk_ids = rank_hybrid(record, alpha)
            metrics = query_metrics(ranked_chunk_ids, record["label_map"], record["ideal_dcg"])
            curve[alpha]["recall_at_10"].append(metrics["recall_at_10"])
            curve[alpha]["ndcg_at_10"].append(metrics["ndcg_at_10"])

    aggregated: Dict[float, Dict[str, Any]] = {}
    for alpha in alpha_values:
        aggregated[alpha] = {
            "recall_at_10": mean_or_zero(curve[alpha]["recall_at_10"]),
            "ndcg_at_10": mean_or_zero(curve[alpha]["ndcg_at_10"]),
            "n_queries": len(records),
        }
    return aggregated


def evaluate_alpha_recall_curve(records: Sequence[Dict[str, Any]], alpha_values: Sequence[float]) -> Dict[float, float]:
    recall_curve: Dict[float, List[float]] = {alpha: [] for alpha in alpha_values}
    for record in records:
        for alpha in alpha_values:
            ranked_chunk_ids = rank_hybrid(record, alpha)
            metrics = query_metrics(ranked_chunk_ids, record["label_map"], record["ideal_dcg"])
            recall_curve[alpha].append(metrics["recall_at_10"])
    return {alpha: mean_or_zero(values) for alpha, values in recall_curve.items()}


def best_alpha_from_curve(curve: Dict[float, Dict[str, Any]], metric_name: str) -> Dict[str, Any]:
    alpha = max(
        curve.keys(),
        key=lambda candidate: (
            curve[candidate][metric_name],
            curve[candidate]["ndcg_at_10"] if metric_name == "recall_at_10" else curve[candidate]["recall_at_10"],
            -candidate,
        ),
    )
    return {
        "alpha": alpha,
        "recall_at_10": curve[alpha]["recall_at_10"],
        "ndcg_at_10": curve[alpha]["ndcg_at_10"],
    }


def evaluate_pilot_hybrid(
    records: Sequence[Dict[str, Any]],
    dataset_name: str,
    seed: int = 2026,
    sample_size: int = PILOT_SAMPLE_SIZE,
    alpha_step: float = 0.05,
    top_alpha_count: int = PILOT_TOP_ALPHA_COUNT,
    min_queries: int = PILOT_MIN_QUERIES,
) -> Dict[str, Any]:
    if len(records) < min_queries:
        return {
            "skipped": True,
            "reason": f"query_count<{min_queries}",
            "query_count": len(records),
        }

    rng = random.Random(stable_seed(seed, dataset_name))
    pilot_pool_indices = sorted(rng.sample(range(len(records)), sample_size))
    pilot_pool = [records[idx] for idx in pilot_pool_indices]
    pilot_pool_index_set = set(pilot_pool_indices)
    holdout_records = [records[idx] for idx in range(len(records)) if idx not in pilot_pool_index_set]

    pilot_alpha_values = build_alpha_values(alpha_step)
    pilot_curve = evaluate_alpha_curve(pilot_pool, pilot_alpha_values)
    ranked_alphas = sorted(
        (
            (alpha, values["recall_at_10"], values["ndcg_at_10"])
            for alpha, values in pilot_curve.items()
        ),
        key=lambda item: (-item[1], -item[2], item[0]),
    )
    selected_alpha_rows = ranked_alphas[:top_alpha_count]
    selected_alphas = [alpha for alpha, _, _ in selected_alpha_rows]

    final_alpha = mean(selected_alphas)
    holdout_metrics = evaluate_alpha_curve(holdout_records, [final_alpha])
    final_metrics = holdout_metrics[final_alpha]

    return {
        "skipped": False,
        "query_count": len(records),
        "pilot_pool_query_count": len(pilot_pool),
        "holdout_query_count": len(holdout_records),
        "pilot_alpha_step": alpha_step,
        "pilot_alpha_values": pilot_alpha_values,
        "pilot_curve": summarize_curve(pilot_curve),
        "selected_alpha_count": top_alpha_count,
        "selected_alpha_rows": [
            {
                "alpha": alpha,
                "recall_at_10": recall_at_10,
                "ndcg_at_10": ndcg_at_10,
            }
            for alpha, recall_at_10, ndcg_at_10 in selected_alpha_rows
        ],
        "sample_size": sample_size,
        "seed": seed,
        "selected_alphas": selected_alphas,
        "final_alpha": final_alpha,
        "tuning_objective": "recall_at_10",
        "metrics": final_metrics,
        "pilot_query_indices": pilot_pool_indices,
        "pilot_queries": [
            {
                "file_name": record["file_name"],
                "query_idx": record["query_idx"],
            }
            for record in pilot_pool
        ],
    }


def summarize_curve(curve: Dict[float, Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "curve": {
            f"{alpha:.2f}": {
                "recall_at_10": values["recall_at_10"],
                "ndcg_at_10": values["ndcg_at_10"],
                "n_queries": values["n_queries"],
            }
            for alpha, values in sorted(curve.items())
        },
        "best_by_recall": best_alpha_from_curve(curve, "recall_at_10"),
        "best_by_ndcg": best_alpha_from_curve(curve, "ndcg_at_10"),
    }


def evaluate_dataset(
    records: Sequence[Dict[str, Any]],
    dataset_name: str,
    seed: int,
    pilot_alpha_step: float,
    pilot_top_alpha_count: int,
) -> Dict[str, Any]:
    fixed_methods = {
        "sparse_only": evaluate_fixed_method(records, "sparse_only"),
        "dense_only": evaluate_fixed_method(records, "dense_only"),
        "rrf": evaluate_fixed_method(records, "rrf"),
    }
    alpha_curve = evaluate_alpha_curve(records, ALPHA_VALUES)
    pilot_hybrid = evaluate_pilot_hybrid(
        records,
        dataset_name=dataset_name,
        seed=seed,
        alpha_step=pilot_alpha_step,
        top_alpha_count=pilot_top_alpha_count,
    )

    return {
        "dataset_name": dataset_name,
        "query_count": len(records),
        "methods": fixed_methods,
        "alpha_sweep": summarize_curve(alpha_curve),
        "pilot_hybrid": pilot_hybrid,
    }


def format_float(value: float) -> str:
    return f"{value:.4f}"


def markdown_table(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> str:
    string_rows = [[str(cell) for cell in row] for row in rows]
    widths = [len(header) for header in headers]
    for row in string_rows:
        for idx, cell in enumerate(row):
            widths[idx] = max(widths[idx], len(cell))

    def format_row(row: Sequence[str]) -> str:
        return "| " + " | ".join(cell.ljust(widths[idx]) for idx, cell in enumerate(row)) + " |"

    header_line = format_row(list(headers))
    separator_line = "| " + " | ".join("-" * widths[idx] for idx in range(len(headers))) + " |"
    body_lines = [format_row(row) for row in string_rows]
    return "\n".join([header_line, separator_line, *body_lines])


def render_dataset_section(result: Dict[str, Any], title_level: int = 2) -> List[str]:
    prefix = "#" * title_level
    lines: List[str] = [f"{prefix} {result['dataset_name']}", ""]
    lines.append(
        f"- Queries: {result['query_count']}"
        + (f" | Pilot holdout: {result['pilot_hybrid'].get('holdout_query_count', 0)}" if not result['pilot_hybrid'].get("skipped") else "")
    )
    if result["pilot_hybrid"].get("skipped"):
        lines.append(f"- Pilot hybrid: skipped ({result['pilot_hybrid'].get('reason', 'n/a')})")
    else:
        selected_alphas = ", ".join(format_float(alpha) for alpha in result["pilot_hybrid"]["selected_alphas"])
        lines.append(
            f"- Pilot hybrid: final alpha={format_float(result['pilot_hybrid']['final_alpha'])}, "
            f"sample_size={result['pilot_hybrid']['sample_size']}, alpha_step={format_float(result['pilot_hybrid']['pilot_alpha_step'])}, "
            f"top_k={result['pilot_hybrid']['selected_alpha_count']}, selected=[{selected_alphas}], "
            f"holdout={result['pilot_hybrid']['holdout_query_count']}, objective={result['pilot_hybrid']['tuning_objective']}"
        )
    lines.append("")

    methods = result["methods"]
    alpha_sweep = result["alpha_sweep"]
    pilot = result["pilot_hybrid"]

    summary_rows = [
        ["Sparse only", format_float(methods["sparse_only"]["recall_at_10"]), format_float(methods["sparse_only"]["ndcg_at_10"]), "BM25 top-10"],
        ["Dense only", format_float(methods["dense_only"]["recall_at_10"]), format_float(methods["dense_only"]["ndcg_at_10"]), "Dense top-10"],
        [
            "Alpha sweep (best recall)",
            format_float(alpha_sweep["best_by_recall"]["recall_at_10"]),
            format_float(alpha_sweep["best_by_recall"]["ndcg_at_10"]),
            f"alpha={format_float(alpha_sweep['best_by_recall']['alpha'])}",
        ],
        [
            "Alpha sweep (best NDCG)",
            format_float(alpha_sweep["best_by_ndcg"]["recall_at_10"]),
            format_float(alpha_sweep["best_by_ndcg"]["ndcg_at_10"]),
            f"alpha={format_float(alpha_sweep['best_by_ndcg']['alpha'])}",
        ],
        ["RRF", format_float(methods["rrf"]["recall_at_10"]), format_float(methods["rrf"]["ndcg_at_10"]), "1/(1+rank), top-10"],
    ]
    if not pilot.get("skipped"):
        summary_rows.append(
            [
                "Pilot hybrid",
                format_float(pilot["metrics"]["recall_at_10"]),
                format_float(pilot["metrics"]["ndcg_at_10"]),
                f"alpha={format_float(pilot['final_alpha'])}",
            ]
        )
    else:
        summary_rows.append(["Pilot hybrid", "SKIPPED", "SKIPPED", pilot.get("reason", "n/a")])

    lines.append("### Method Summary")
    lines.append("")
    lines.append(markdown_table(["Method", "Recall@10", "NDCG@10", "Notes"], summary_rows))
    lines.append("")

    alpha_rows = []
    for alpha_str, values in alpha_sweep["curve"].items():
        alpha_rows.append(
            [
                alpha_str,
                format_float(values["recall_at_10"]),
                format_float(values["ndcg_at_10"]),
            ]
        )
    lines.append("### Alpha Sweep Curve")
    lines.append("")
    lines.append(markdown_table(["Alpha", "Recall@10", "NDCG@10"], alpha_rows))
    lines.append("")

    if not pilot.get("skipped"):
        pilot_rows = [
            [idx + 1, format_float(row["alpha"]), format_float(row["recall_at_10"]), format_float(row["ndcg_at_10"])]
            for idx, row in enumerate(pilot["selected_alpha_rows"])
        ]
        lines.append("### Pilot Hybrid Details")
        lines.append("")
        lines.append(markdown_table(["Rank", "Alpha", "Recall@10", "NDCG@10"], pilot_rows))
        lines.append("")

    return lines


def build_report(results: Dict[str, Any]) -> List[str]:
    lines: List[str] = [
        "# Hybrid Retrieval Benchmark Report",
        "",
        "## Methodology",
        "",
        "- Recall@10 is computed as relevant_count / 10 for the retrieved top-10 results.",
        "- NDCG@10 uses binary relevance labels from the judge caches.",
        "- Hybrid scores are built from the top-10 dense and sparse scores after per-source min-max normalization.",
        "- RRF uses 1 / (1 + rank) for ranks 1 to 10 from each retriever.",
        "- Pilot hybrid uses a 50-query pilot pool, evaluates alpha values from 0.00 to 1.00 in 0.05 steps, and sets the final alpha to the mean of the top 3 alphas by Recall@10 on the pilot pool.",
        "- Pilot hybrid is skipped for datasets with fewer than 100 queries.",
        "- The front dataset section shows AIHub topics merged by TL+VL pair; KorMedMCQA appears only in the group-level section.",
        "",
    ]

    lines.extend(["## Dataset-Level Analysis", ""])
    for dataset_name in sorted(results["aihub_datasets"].keys()):
        dataset_result = results["aihub_datasets"][dataset_name]
        dataset_result = dict(dataset_result)
        dataset_result["dataset_name"] = f"{dataset_name}(TL+VL)"
        lines.extend(render_dataset_section(dataset_result, title_level=3))

    lines.extend(["## Group-Level Analysis", ""])
    for group_name in ["AIHub TL+VL", "KorMedMCQA_dentist", "KorMedMCQA_doctor"]:
        lines.extend(render_dataset_section(results["groups"][group_name], title_level=3))

    lines.extend(
        [
            "## Summary",
            "",
            "- Full alpha curves and per-dataset details are also included in the JSON companion file.",
            "- File-level pooling is query-weighted; group-level results pool all queries belonging to the group.",
            "",
        ]
    )
    return lines


def build_results_payload(
    file_results: Dict[str, Dict[str, Any]],
    aihub_dataset_results: Dict[str, Dict[str, Any]],
    group_results: Dict[str, Dict[str, Any]],
    config: Dict[str, Any],
) -> Dict[str, Any]:
    return {
        "generated_at": datetime.now().isoformat(),
        "config": config,
        "files": file_results,
        "aihub_datasets": aihub_dataset_results,
        "groups": group_results,
    }



def safe_sheet_name(name: str, used: Optional[set[str]] = None) -> str:
    invalid = set('[]:*?/\\')
    cleaned = ''.join('_' if ch in invalid else ch for ch in str(name)).strip() or 'Sheet'
    cleaned = cleaned[:31]
    if used is None:
        return cleaned
    candidate = cleaned
    suffix = 1
    while candidate in used:
        suffix_text = f"_{suffix}"
        candidate = cleaned[: 31 - len(suffix_text)] + suffix_text
        suffix += 1
    used.add(candidate)
    return candidate


def autosize_worksheet(ws: Any, max_width: int = 45) -> None:
    from openpyxl.utils import get_column_letter

    for column_cells in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), max_width)


def write_rows(ws: Any, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))


def style_table(ws: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if ws.max_row < 1 or ws.max_column < 1:
        return
    header_fill = PatternFill('solid', fgColor='EAF2F8')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=False)
    if ws.max_row >= 2:
        table_name = ''.join(ch for ch in ws.title if ch.isalnum())[:20] or 'Table'
        table = Table(displayName=f"{table_name}Table", ref=ws.dimensions)
        style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    ws.freeze_panes = 'A2'
    autosize_worksheet(ws)


def method_summary_rows(result: Dict[str, Any], scope: str, name: str) -> List[List[Any]]:
    rows: List[List[Any]] = []
    methods = result['methods']
    for method_key, label in [('sparse_only', 'Sparse only'), ('dense_only', 'Dense only'), ('rrf', 'RRF')]:
        rows.append([
            scope,
            name,
            label,
            methods[method_key]['recall_at_10'],
            methods[method_key]['ndcg_at_10'],
            methods[method_key]['n_queries'],
            '',
            '',
            '',
        ])

    best_recall = result['alpha_sweep']['best_by_recall']
    rows.append([
        scope,
        name,
        'Alpha sweep (best recall)',
        best_recall['recall_at_10'],
        best_recall['ndcg_at_10'],
        result['query_count'],
        best_recall['alpha'],
        '',
        '',
    ])

    best_ndcg = result['alpha_sweep']['best_by_ndcg']
    rows.append([
        scope,
        name,
        'Alpha sweep (best NDCG)',
        best_ndcg['recall_at_10'],
        best_ndcg['ndcg_at_10'],
        result['query_count'],
        best_ndcg['alpha'],
        '',
        '',
    ])

    pilot = result['pilot_hybrid']
    if pilot.get('skipped'):
        rows.append([scope, name, 'Pilot hybrid', '', '', result['query_count'], '', 'SKIPPED', pilot.get('reason', '')])
    else:
        rows.append([
            scope,
            name,
            'Pilot hybrid',
            pilot['metrics']['recall_at_10'],
            pilot['metrics']['ndcg_at_10'],
            pilot['metrics']['n_queries'],
            pilot['final_alpha'],
            'OK',
            f"selected={pilot['selected_alphas']}",
        ])
    return rows


def alpha_curve_rows(result: Dict[str, Any], scope: str, name: str) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for alpha, values in result['alpha_sweep']['curve'].items():
        rows.append([
            scope,
            name,
            float(alpha),
            values['recall_at_10'],
            values['ndcg_at_10'],
            values['n_queries'],
        ])
    return rows


def pilot_detail_rows(result: Dict[str, Any], scope: str, name: str) -> List[List[Any]]:
    pilot = result['pilot_hybrid']
    if pilot.get('skipped'):
        return [[scope, name, '', '', '', '', 'SKIPPED', pilot.get('reason', '')]]
    return [
        [
            scope,
            name,
            idx + 1,
            row['alpha'],
            row['recall_at_10'],
            row['ndcg_at_10'],
            pilot['final_alpha'],
            '',
        ]
        for idx, row in enumerate(pilot['selected_alpha_rows'])
    ]


def save_excel_report(path: Path, payload: Dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError('Excel output requires openpyxl. Install it with: pip install openpyxl') from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Method Summary'

    method_headers = ['Scope', 'Dataset', 'Method', 'Recall@10', 'NDCG@10', 'N Queries', 'Alpha', 'Status', 'Notes']
    method_rows: List[List[Any]] = []
    curve_rows: List[List[Any]] = []
    pilot_rows: List[List[Any]] = []

    for file_name, result in sorted(payload['files'].items()):
        method_rows.extend(method_summary_rows(result, 'file', file_name))
        curve_rows.extend(alpha_curve_rows(result, 'file', file_name))
        pilot_rows.extend(pilot_detail_rows(result, 'file', file_name))

    for dataset_name, result in sorted(payload['aihub_datasets'].items()):
        method_rows.extend(method_summary_rows(result, 'aihub_dataset', dataset_name))
        curve_rows.extend(alpha_curve_rows(result, 'aihub_dataset', dataset_name))
        pilot_rows.extend(pilot_detail_rows(result, 'aihub_dataset', dataset_name))

    for group_name, result in sorted(payload['groups'].items()):
        method_rows.extend(method_summary_rows(result, 'group', group_name))
        curve_rows.extend(alpha_curve_rows(result, 'group', group_name))
        pilot_rows.extend(pilot_detail_rows(result, 'group', group_name))

    write_rows(ws, method_headers, method_rows)
    style_table(ws)

    ws_curve = wb.create_sheet('Alpha Curves')
    write_rows(ws_curve, ['Scope', 'Dataset', 'Alpha', 'Recall@10', 'NDCG@10', 'N Queries'], curve_rows)
    style_table(ws_curve)

    ws_pilot = wb.create_sheet('Pilot Details')
    write_rows(ws_pilot, ['Scope', 'Dataset', 'Rank', 'Alpha', 'Recall@10', 'NDCG@10', 'Final Alpha', 'Notes'], pilot_rows)
    style_table(ws_pilot)

    ws_config = wb.create_sheet('Config')
    config_rows = [[key, json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value] for key, value in payload['config'].items()]
    config_rows.insert(0, ['generated_at', payload.get('generated_at', '')])
    write_rows(ws_config, ['Key', 'Value'], config_rows)
    style_table(ws_config)

    for ws_item in wb.worksheets:
        for row in ws_item.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'
    wb.save(path)

def save_pilot_selected_queries(path: Path, payload: Dict[str, Any]) -> None:
    rows: List[Dict[str, Any]] = []

    for scope, result_dict in [
        ("file", payload["files"]),
        ("aihub_dataset", payload["aihub_datasets"]),
        ("group", payload["groups"]),
    ]:
        for dataset_name, result in sorted(result_dict.items()):
            pilot = result["pilot_hybrid"]

            if pilot.get("skipped"):
                rows.append({
                    "scope": scope,
                    "dataset": dataset_name,
                    "status": "SKIPPED",
                    "reason": pilot.get("reason", ""),
                    "pilot_queries": [],
                })
                continue

            rows.append({
                "scope": scope,
                "dataset": dataset_name,
                "status": "OK",
                "sample_size": pilot["sample_size"],
                "seed": pilot["seed"],
                "pilot_pool_query_count": pilot["pilot_pool_query_count"],
                "holdout_query_count": pilot["holdout_query_count"],
                "pilot_query_indices": pilot["pilot_query_indices"],
                "pilot_queries": pilot["pilot_queries"],
            })

    save_json(path, rows)

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a multi-level hybrid retrieval benchmark report")
    parser.add_argument("--cache-root", type=str, default="retrieval_cache")
    parser.add_argument("--report-md", type=str, default=DEFAULT_REPORT_MD)
    parser.add_argument("--report-json", type=str, default=DEFAULT_REPORT_JSON)
    parser.add_argument("--report-xlsx", type=str, default=DEFAULT_REPORT_XLSX)
    parser.add_argument("--include-files", type=str, default="", help="Comma-separated file names to evaluate")
    parser.add_argument("--max-files", type=int, default=0, help="0 means all files")
    parser.add_argument("--seed", type=int, default=2026)
    parser.add_argument("--alpha-step", type=float, default=0.1)
    parser.add_argument("--pilot-alpha-step", type=float, default=0.05)
    parser.add_argument("--pilot-sample-size", type=int, default=PILOT_SAMPLE_SIZE)
    parser.add_argument(
        "--pilot-top-alpha-count",
        "--pilot-bootstrap-repeats",
        dest="pilot_top_alpha_count",
        type=int,
        default=PILOT_TOP_ALPHA_COUNT,
        help="Number of top pilot alphas to average (old flag kept for compatibility)",
    )
    parser.add_argument("--pilot-min-queries", type=int, default=PILOT_MIN_QUERIES)
    parser.add_argument("--max-queries-per-file", type=int, default=0, help="0 means all queries per file")
    args = parser.parse_args()

    if args.alpha_step <= 0:
        raise ValueError("--alpha-step must be > 0")
    if args.pilot_alpha_step <= 0:
        raise ValueError("--pilot-alpha-step must be > 0")
    if args.pilot_sample_size < 1:
        raise ValueError("--pilot-sample-size must be >= 1")
    if args.pilot_top_alpha_count < 1:
        raise ValueError("--pilot-top-alpha-count must be >= 1")
    if args.pilot_min_queries < 1:
        raise ValueError("--pilot-min-queries must be >= 1")

    workspace_root = Path(__file__).resolve().parent
    cache_root = (workspace_root / args.cache_root).resolve()
    dense_root = cache_root / "dense"
    bm25_root = cache_root / "bm25"
    dense_judge_root = cache_root / "dense_llmjudge"
    bm25_judge_root = cache_root / "bm25_llmjudge"

    report_md_path = (workspace_root / args.report_md).resolve()
    report_json_path = (workspace_root / args.report_json).resolve()
    report_xlsx_path = (workspace_root / args.report_xlsx).resolve()
    pilot_selected_query_path = (workspace_root / "eval_result_ret/pilot_selected_query.json").resolve()
    report_md_path.parent.mkdir(parents=True, exist_ok=True)
    report_json_path.parent.mkdir(parents=True, exist_ok=True)
    report_xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    file_names = discover_cache_files(dense_root)
    include_files = [name.strip() for name in args.include_files.split(",") if name.strip()]
    file_names = filter_file_names(file_names, include_files=include_files, max_files=args.max_files)

    if not file_names:
        raise ValueError("No cache files selected for evaluation")

    print(f"Selected {len(file_names)} files from {dense_root}")

    file_results: Dict[str, Dict[str, Any]] = {}
    aihub_dataset_records: Dict[str, List[Dict[str, Any]]] = {}
    group_records: Dict[str, List[Dict[str, Any]]] = {
        "AIHub TL+VL": [],
        "KorMedMCQA_dentist": [],
        "KorMedMCQA_doctor": [],
    }

    for file_name in file_names:
        dense_cache = load_retrieval_results(dense_root, file_name)
        bm25_cache = load_retrieval_results(bm25_root, file_name)
        dense_judge_cache = load_retrieval_results(dense_judge_root, file_name)
        bm25_judge_cache = load_retrieval_results(bm25_judge_root, file_name)

        records = build_query_records_for_file(
            file_name=file_name,
            dense_cache=dense_cache,
            bm25_cache=bm25_cache,
            dense_judge_cache=dense_judge_cache,
            bm25_judge_cache=bm25_judge_cache,
        )
        if args.max_queries_per_file > 0:
            records = records[: args.max_queries_per_file]

        dataset_result = evaluate_dataset(
            records,
            dataset_name=file_name,
            seed=args.seed,
            pilot_alpha_step=args.pilot_alpha_step,
            pilot_top_alpha_count=args.pilot_top_alpha_count,
        )
        file_results[file_name] = dataset_result

        group_name = group_name_for_file(file_name)
        if group_name is not None:
            group_records[group_name].extend(records)

        aihub_dataset_name = aihub_dataset_name_for_file(file_name)
        if aihub_dataset_name is not None:
            aihub_dataset_records.setdefault(aihub_dataset_name, []).extend(records)

        print(
            f"Completed {file_name}: queries={dataset_result['query_count']} "
            f"sparse_recall={dataset_result['methods']['sparse_only']['recall_at_10']:.4f} "
            f"dense_recall={dataset_result['methods']['dense_only']['recall_at_10']:.4f}"
        )

    aihub_dataset_results: Dict[str, Dict[str, Any]] = {}
    for dataset_name, records in aihub_dataset_records.items():
        aihub_dataset_results[dataset_name] = evaluate_dataset(
            records,
            dataset_name=f"{dataset_name}(TL+VL)",
            seed=args.seed,
            pilot_alpha_step=args.pilot_alpha_step,
            pilot_top_alpha_count=args.pilot_top_alpha_count,
        )

    group_results: Dict[str, Dict[str, Any]] = {}
    for group_name, records in group_records.items():
        group_results[group_name] = evaluate_dataset(
            records,
            dataset_name=group_name,
            seed=args.seed,
            pilot_alpha_step=args.pilot_alpha_step,
            pilot_top_alpha_count=args.pilot_top_alpha_count,
        )

    payload = build_results_payload(
        file_results=file_results,
        aihub_dataset_results=aihub_dataset_results,
        group_results=group_results,
        config={
            "cache_root": str(cache_root),
            "dense_root": str(dense_root),
            "bm25_root": str(bm25_root),
            "dense_judge_root": str(dense_judge_root),
            "bm25_judge_root": str(bm25_judge_root),
            "alpha_step": args.alpha_step,
            "pilot_alpha_step": args.pilot_alpha_step,
            "pilot_sample_size": args.pilot_sample_size,
            "pilot_top_alpha_count": args.pilot_top_alpha_count,
            "pilot_min_queries": args.pilot_min_queries,
            "seed": args.seed,
            "selected_files": file_names,
            "report_xlsx": str(report_xlsx_path),
        },
    )

    report_lines = build_report(payload)
    report_md_path.write_text("\n".join(report_lines), encoding="utf-8")
    save_json(report_json_path, payload)
    save_pilot_selected_queries(pilot_selected_query_path, payload)
    save_excel_report(report_xlsx_path, payload)

    print(f"Saved markdown report: {report_md_path}")
    print(f"Saved JSON summary: {report_json_path}")
    print(f"Saved Excel summary: {report_xlsx_path}")
    print(f"Saved pilot selected queries: {pilot_selected_query_path}")


if __name__ == "__main__":
    main()
